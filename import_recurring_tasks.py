"""
Import recurring tasks from Excel sheet "Họp hàng tuần"
into recurring_task_templates table.
Run from project root: python import_recurring_tasks.py
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from sqlalchemy.orm import Session
from database import SessionLocal
from models import User, RecurringTaskTemplate

EXCEL_PATH = r'E:\Downloads\Ban KSNB - To do list.xlsx'
SHEET_NAME = 'Họp hàng tuần'

FREQ_MAP = {
    9:  'weekly',
    10: 'monthly',
    11: 'quarterly',
    12: 'semi_annual',
    13: 'annual',
    14: 'ad_hoc',
}

def parse_tasks(cell_value: str) -> list[str]:
    """Split numbered list into individual task titles (first line only)."""
    if not cell_value or not str(cell_value).strip():
        return []
    text = str(cell_value).strip()
    # Split on numbered items: "1. " "2. " etc.
    parts = re.split(r'\n?(?<!\d)\d{1,2}\.\s+', text)
    tasks = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Take only first line (before sub-bullets or extra newlines)
        first_line = part.split('\n')[0].strip().rstrip('.')
        if first_line:
            tasks.append(first_line)
    return tasks


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Collect rows: col 1 = name, freq cols defined in FREQ_MAP
    person_rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        name = row[1]
        if not name or not str(name).strip():
            continue
        entry = {'name': str(name).strip(), 'tasks': {}}
        for col_idx, freq_key in FREQ_MAP.items():
            entry['tasks'][freq_key] = parse_tasks(row[col_idx])
        person_rows.append(entry)

    db: Session = SessionLocal()
    try:
        # Load all users
        all_users = db.query(User).filter(User.is_active == True).all()
        user_map = {u.full_name.strip(): u for u in all_users if u.full_name}

        inserted = 0
        skipped_names = []

        for entry in person_rows:
            name = entry['name']
            user = user_map.get(name)
            if not user:
                # Try partial match (last name + first name)
                matches = [u for fn, u in user_map.items()
                           if name.lower() in fn.lower() or fn.lower() in name.lower()]
                if len(matches) == 1:
                    user = matches[0]
                    print(f'  ~ Matched "{name}" → "{user.full_name}"')
                else:
                    skipped_names.append(name)
                    print(f'  ✗ Not found: "{name}"')
                    continue

            # Check existing tasks for this user to avoid duplicates
            existing = db.query(RecurringTaskTemplate)\
                .filter(RecurringTaskTemplate.user_id == user.id).all()
            existing_titles = {t.title.strip().lower() for t in existing}

            for freq, titles in entry['tasks'].items():
                for title in titles:
                    if title.lower() in existing_titles:
                        print(f'    skip (exists): {title}')
                        continue
                    t = RecurringTaskTemplate(
                        user_id=user.id,
                        title=title,
                        frequency=freq,
                        is_active=True,
                    )
                    db.add(t)
                    existing_titles.add(title.lower())
                    inserted += 1
                    print(f'    + [{freq}] {title}')

        db.commit()
        print(f'\n✓ Done. Inserted {inserted} tasks.')
        if skipped_names:
            print(f'  Not matched: {skipped_names}')
    finally:
        db.close()


if __name__ == '__main__':
    main()
