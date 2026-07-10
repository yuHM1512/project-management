"""Fix: bad 'trục' entries + re-import 2 unmatched users with accent-normalized matching."""
import sys, io, re, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from database import SessionLocal
from models import User, RecurringTaskTemplate

EXCEL_PATH = r'E:\Downloads\Ban KSNB - To do list.xlsx'
SHEET_NAME = 'Họp hàng tuần'
FREQ_MAP = {9:'weekly', 10:'monthly', 11:'quarterly', 12:'semi_annual', 13:'annual', 14:'ad_hoc'}

# Correct titles for "trục" tasks (col index → correct title)
TRUC_FIX = {
    14: 'Triển khai hành động OPEX trục 4',  # Phạm Ngọc Minh Trí
    13: 'Triển khai hành động OPEX trục 3',  # Lê Thị Kim Ngân
    12: 'Triển khai hành động OPEX trục 2',  # Nguyễn Thị Ngọc Hoa
    11: 'Triển khai hành động OPEX trục 1',  # Nguyễn Khắc Minh Huy
}

def strip_accents(s: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFD', s.lower())
        if unicodedata.category(c) != 'Mn'
    )

def parse_tasks(cell_value) -> list[str]:
    """Only split on lines that START with a number-period (e.g. '1. Task')."""
    if not cell_value or not str(cell_value).strip():
        return []
    lines = str(cell_value).strip().split('\n')
    tasks, current = [], None
    for line in lines:
        m = re.match(r'^\d{1,2}\.\s+(.+)', line.strip())
        if m:
            if current is not None:
                tasks.append(current)
            current = m.group(1).strip().rstrip('.')
    if current is not None:
        tasks.append(current)
    return tasks

def main():
    db = SessionLocal()
    try:
        # ── 1. Fix bad 'trục' tasks ──────────────────────────────────────
        bad = db.query(RecurringTaskTemplate).filter(
            RecurringTaskTemplate.title == 'Triển khai hành động OPEX trục'
        ).all()
        for t in bad:
            correct = TRUC_FIX.get(t.user_id)
            if correct:
                print(f'Fix user {t.user_id}: "{t.title}" → "{correct}"')
                t.title = correct
        db.commit()

        # ── 2. Import 2 missing users ─────────────────────────────────────
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb[SHEET_NAME]

        all_users = db.query(User).filter(User.is_active == True).all()
        # Build lookup: stripped-accent full_name → user
        user_map = {strip_accents(u.full_name): u for u in all_users if u.full_name}

        target_names = ['Huỳnh Thị Yến Linh', 'Ngô Văn Thắng']

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            raw_name = row[1]
            if not raw_name or str(raw_name).strip() not in target_names:
                continue
            name = str(raw_name).strip()
            key = strip_accents(name)
            user = user_map.get(key)
            if not user:
                print(f'Still not found: {name} (key={key})')
                print('Available keys:', [k for k in user_map if 'huynh' in k or 'ngo' in k])
                continue

            print(f'\nImporting for: {name} → DB: {user.full_name} (id={user.id})')
            existing = db.query(RecurringTaskTemplate).filter(
                RecurringTaskTemplate.user_id == user.id).all()
            existing_titles = {t.title.strip().lower() for t in existing}

            inserted = 0
            for col_idx, freq_key in FREQ_MAP.items():
                for title in parse_tasks(row[col_idx]):
                    if title.lower() in existing_titles:
                        print(f'  skip: {title}')
                        continue
                    db.add(RecurringTaskTemplate(
                        user_id=user.id, title=title,
                        frequency=freq_key, is_active=True))
                    existing_titles.add(title.lower())
                    inserted += 1
                    print(f'  + [{freq_key}] {title}')

        db.commit()
        print('\n✓ Fix + import done.')
    finally:
        db.close()

if __name__ == '__main__':
    main()
