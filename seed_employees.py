#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script để seed dữ liệu nhân viên từ file Excel vào table users
Usage: python seed_employees.py
"""

import os
import sys
import io
import math
import unicodedata
from pathlib import Path
import openpyxl
from passlib.context import CryptContext
from datetime import datetime

# Fix encoding for Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from database import SessionLocal, engine
from models import User

# Hash password context
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def normalize_text(text: str) -> str:
    """Normalize text by removing accents"""
    if not text:
        return ""
    # Remove diacritics
    nfd_form = unicodedata.normalize('NFD', text)
    return ''.join(c for c in nfd_form if unicodedata.category(c) != 'Mn').lower()

def get_password_hash(password: str) -> str:
    """Hash mật khẩu"""
    return pwd_context.hash(password)


def cell_to_text(value) -> str:
    """Normalize worksheet cell values into trimmed text."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def parse_field_values(raw_value) -> list[str]:
    """Split one field cell into unique field values."""
    raw_text = cell_to_text(raw_value)
    if not raw_text:
        return []

    fields = []
    for part in raw_text.replace(";", ",").split(","):
        field_name = part.strip().upper()
        if field_name and field_name not in fields:
            fields.append(field_name)
    return fields


def parse_chapter_values(raw_value) -> list[str]:
    """Split one chapter cell into a clean list of chapters."""
    raw_text = cell_to_text(raw_value)
    if not raw_text:
        return []

    chapters = []
    for part in raw_text.replace(";", ",").split(","):
        chapter_name = part.strip()
        if chapter_name and chapter_name not in chapters:
            chapters.append(chapter_name)
    return chapters


def parse_group_value(raw_value) -> str:
    """Convert values like 'Nhóm 1' into '1'."""
    raw_text = cell_to_text(raw_value)
    if not raw_text:
        return ""

    normalized = raw_text.lower().replace("nhóm", "").replace("nhom", "").strip()
    return normalized or raw_text


def build_user_metadata(field_raw, chapter_raw, group_raw):
    """Build users.field, users.chapter, and users.group payloads."""
    fields = parse_field_values(field_raw)
    chapters = parse_chapter_values(chapter_raw)
    group_value = parse_group_value(group_raw)

    chapter_entries = []
    group_entries = []
    for field_name in fields:
        chapter_entries.append({
            "field": field_name,
            "chapters": chapters,
        })
        if group_value:
            group_entries.append({
                "field": field_name,
                "group": group_value,
            })

    return fields, chapter_entries, group_entries

def seed_employees(domain: str = "company.com"):
    """Đọc file Excel và insert nhân viên vào database"""
    excel_file = Path("seed_employees.xlsx")

    if not excel_file.exists():
        print(f"[ERROR] File khong ton tai: {excel_file}")
        sys.exit(1)

    # Doc file Excel
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        print(f"[OK] Mo file Excel: {excel_file}")
    except Exception as e:
        print(f"[ERROR] Loi doc file Excel: {e}")
        sys.exit(1)

    # Lay headers tu dong dau tien
    headers = {}
    header_names = []  # Keep track of header names for easier access
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(1, col_idx)
        if cell.value:
            header_key = str(cell.value).lower().strip()
            headers[header_key] = col_idx
            header_names.append((header_key, col_idx))

    print(f"\n[INFO] Headers tim thay:")
    for k, idx in header_names:
        print(f"  - {k} -> Col {idx}")

    # Map columns from Excel to User model
    # Expected columns: "Ho & ten", "Ma nhan vien", "Don vi", "Chuc danh", "Linh vuc", "Chuong", "Nhom"
    col_mapping = {
        'ho & tên': 'full_name',
        'họ & tên': 'full_name',
        'mã nhân viên': 'username',
        'đơn vị': 'department',
        'chức danh': 'position',
        'lĩnh vực': 'field',
        'chương': 'chapter',
        'nhóm': 'group',
    }

    required_source_fields = ['họ & tên', 'mã nhân viên']
    missing_fields = [f for f in required_source_fields if f not in headers and 'ho & tên' not in headers]

    if missing_fields and 'ho & tên' not in headers:
        print(f"[ERROR] Thieu cac cot bat buoc: {missing_fields}")
        sys.exit(1)

    # Session database
    db = SessionLocal()

    try:
        users_created = 0
        users_updated = 0
        users_skipped = 0
        errors = []

        # Xu ly tung dong du lieu
        for row_idx in range(2, worksheet.max_row + 1):
            try:
                # Lay gia tri tu cac cot
                full_name = None
                username = None
                department = None
                position = None
                field_data = None
                chapter_data = None
                group_data = None

                # Find columns by header content
                for key, col_idx in headers.items():
                    cell_value = worksheet.cell(row_idx, col_idx).value
                    # Normalize key for comparison - remove accents and spaces
                    key_norm = normalize_text(key).replace('&', '').replace(' ', '')

                    if 'ho' in key_norm and 'ten' in key_norm:
                        full_name = cell_value
                    elif 'ma' in key_norm and 'nhan' in key_norm and 'vien' in key_norm:
                        username = cell_value
                    elif 'donvi' in key_norm or 'don' in key_norm and 'vi' in key_norm:
                        department = cell_value
                    elif 'chucdanh' in key_norm or 'chuc' in key_norm and 'danh' in key_norm:
                        position = cell_value
                    elif 'linhvuc' in key_norm or 'linh' in key_norm and 'vuc' in key_norm:
                        field_data = cell_value
                    elif 'chuong' in key_norm:
                        chapter_data = cell_value
                    elif 'nhom' in key_norm:
                        group_data = cell_value

                # Skip empty rows
                if not username or not full_name:
                    continue

                username = cell_to_text(username)
                full_name = cell_to_text(full_name)

                # Convert values to strings
                department = cell_to_text(department) or None
                position = cell_to_text(position) or None
                fields, chapter_entries, group_entries = build_user_metadata(
                    field_data,
                    chapter_data,
                    group_data,
                )

                # Create email from username
                email = f"{username.lower()}@{domain}"

                # Validate
                if not username or not full_name:
                    errors.append(f"Row {row_idx}: Missing username or full_name")
                    users_skipped += 1
                    continue

                # Upsert duplicate
                existing_user = db.query(User).filter(
                    (User.username == username) | (User.email == email)
                ).first()

                if existing_user:
                    existing_user.full_name = full_name
                    existing_user.department = department
                    existing_user.position = position
                    existing_user.field = fields
                    existing_user.chapter = chapter_entries
                    existing_user.group = group_entries
                    existing_user.is_active = True
                    users_updated += 1
                    print(f"[UPDATE] Row {row_idx}: Cap nhat user '{username}' ({email})")
                    continue

                # Tao user moi
                hashed_password = get_password_hash("123456")

                user = User(
                    username=username,
                    email=email,
                    full_name=full_name,
                    hashed_password=hashed_password,
                    department=department,
                    position=position,
                    field=fields,
                    chapter=chapter_entries,
                    group=group_entries,
                    role='member',
                    is_active=True
                )

                db.add(user)
                users_created += 1
                print(f"[OK] Row {row_idx}: Tao user '{username}' ({email})")

            except Exception as e:
                error_msg = f"Row {row_idx}: {str(e)}"
                errors.append(error_msg)
                users_skipped += 1
                print(f"[ERROR] {error_msg}")
                continue

        # Commit changes
        if users_created > 0 or users_updated > 0:
            db.commit()
            print(f"\n[SUCCESS] Tao {users_created} user(s)")
        else:
            print(f"\n[WARN] Khong co user nao duoc tao")

        if users_updated > 0:
            print(f"[SUCCESS] Cap nhat {users_updated} user(s)")

        if users_skipped > 0:
            print(f"[WARN] Skip {users_skipped} row(s)")

        if errors:
            print(f"\n[INFO] Errors:")
            for error in errors[:5]:
                print(f"  - {error}")
            if len(errors) > 5:
                print(f"  ... va {len(errors) - 5} error(s) khac")

    except Exception as e:
        db.rollback()
        print(f"\n[FATAL] Loi khi thuc hien: {e}")
        sys.exit(1)
    finally:
        db.close()

if __name__ == "__main__":
    print("[INFO] Bat dau seed du lieu nhan vien...")
    import sys
    domain = sys.argv[1] if len(sys.argv) > 1 else "company.com"
    seed_employees(domain=domain)
    print("\n[INFO] Hoan thanh!\n")
