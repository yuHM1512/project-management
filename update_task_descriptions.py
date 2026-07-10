"""Update existing recurring tasks with description from Excel bullet points.
Also imports tasks for Yến Linh & Ngô Văn Thắng (no numbered items → treat as bullet-only tasks).
"""
import sys, io, re, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from database import SessionLocal
from models import User, RecurringTaskTemplate

EXCEL_PATH = r'E:\Downloads\Ban KSNB - To do list.xlsx'
SHEET_NAME = 'Họp hàng tuần'
FREQ_MAP = {9:'weekly', 10:'monthly', 11:'quarterly', 12:'semi_annual', 13:'annual', 14:'ad_hoc'}


def strip_accents(s: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFD', s.lower())
        if unicodedata.category(c) != 'Mn'
    )


def parse_tasks_with_desc(cell_value) -> list[dict]:
    """Return list of {title, description} dicts.
    Title = first line of numbered item, description = subsequent bullet lines.
    If no numbered items, treat entire content as description for a single generic task.
    """
    if not cell_value or not str(cell_value).strip():
        return []
    lines = str(cell_value).strip().split('\n')
    results, current_title, current_desc_lines = [], None, []

    for line in lines:
        m = re.match(r'^\d{1,2}\.\s+(.+)', line.strip())
        if m:
            if current_title is not None:
                desc = '\n'.join(l for l in current_desc_lines if l.strip())
                results.append({'title': current_title, 'description': desc or None})
            current_title = m.group(1).strip().rstrip('.')
            current_desc_lines = []
        else:
            stripped = line.strip()
            if stripped and current_title is not None:
                current_desc_lines.append(stripped)

    if current_title is not None:
        desc = '\n'.join(l for l in current_desc_lines if l.strip())
        results.append({'title': current_title, 'description': desc or None})

    return results


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    db = SessionLocal()
    try:
        all_users = db.query(User).filter(User.is_active == True).all()
        user_map_accent = {strip_accents(u.full_name): u for u in all_users if u.full_name}
        user_map_exact  = {u.full_name.strip(): u for u in all_users if u.full_name}

        updated = 0
        inserted = 0

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            raw_name = row[1]
            if not raw_name or not str(raw_name).strip():
                continue
            name = str(raw_name).strip()

            user = user_map_exact.get(name) or user_map_accent.get(strip_accents(name))
            if not user:
                print(f'✗ No match: {name}')
                continue

            # Load existing tasks for this user
            existing = db.query(RecurringTaskTemplate)\
                .filter(RecurringTaskTemplate.user_id == user.id).all()
            title_to_task = {t.title.strip().lower(): t for t in existing}

            for col_idx, freq_key in FREQ_MAP.items():
                for item in parse_tasks_with_desc(row[col_idx]):
                    title = item['title']
                    desc  = item['description']
                    key   = title.strip().lower()
                    task  = title_to_task.get(key)

                    if task:
                        # Update description if changed
                        if task.description != desc:
                            task.description = desc
                            updated += 1
                            print(f'  ~ update desc [{freq_key}] {title[:60]}')
                    else:
                        # Insert new (shouldn't happen for main 5 users)
                        new_t = RecurringTaskTemplate(
                            user_id=user.id, title=title,
                            description=desc, frequency=freq_key, is_active=True)
                        db.add(new_t)
                        title_to_task[key] = new_t
                        inserted += 1
                        print(f'  + insert [{freq_key}] {title[:60]}')

        db.commit()
        print(f'\n✓ Done. Updated={updated} descriptions, Inserted={inserted} new tasks.')
    finally:
        db.close()


if __name__ == '__main__':
    main()
